#!/Library/Frameworks/Python.framework/Versions/2.7/bin/python
from pyexcel_xlsx import save_data
import pyexcel as pe
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import worksheet
import datetime
from dateutil.relativedelta import relativedelta
from datetime import datetime, date, timedelta
import os
import sys
reload(sys) #to help with seemingly random'ascii' encoding error
sys.setdefaultencoding('utf8') # ^^ <--Pythong interpreter doesn't like it, but it works
from collections import OrderedDict
import pandas as pd
from pandas import read_excel
from pyexcel._compact import OrderedDict
import xlsxwriter
import PySimpleGUI27 as sg
from PySimpleGUI27 import SetOptions

## Run this script in terminal `python 'name'.py`
## Everything is setup for Python version 2.7
## If using different version of python, may need to update packages/commands

## This script will open up an excel file, write the headers, and read the file
## First, will return the projects which are active
## --Choose by entering the corresponding number
## Will be followed by confirmation (pretty much everything is)
## --To confirm, enter 'y', or just hit 'Return'
## Once a project is selected (or input), all of the project notes will be listed in the terminal window
## Next, will return the tasks which are active
## -- task selection is same process as project selection
## NOTE: If type in a project or task name that is not active, this will re-activate the project/task
## Once a task is selected (or added), all of the task notes will be listed in the terminal window
## Also, a window will popup where project and task status can be updated, and notes can be added
## When the window pops up, a timer will also start in the top bar (if the application 'Thyme' is installed)
## If want to switch project or task, just click those selections when done
## If you intended to change the project or task status, be sure to do that first!!
## If you are done with the session, click Dunzo

## No matter the selection, after the popup window is closed, if Box is connected, the system will automatically compare the documents and backup any new items to the backup file.
## NOTE: This does mean that if running on separate computers, the logfiles on each computer coudl end up being slightly different.. but the backup will always hold all of the information.

## After backing everything up, the loop will continue as specified, or will terminate and close everything out.


#=======================
# local file parameters
#=======================

local_folder    = '/Users/zjcole/Documents/file_drawer/dev/'
local_filename  = 'zptz_local.xlsx'
computer_name   = 'Donbot' #update this for each computer you use it on
other_computer  = 'Oogway' #update this for each computer you use it on -- should be the name of the inactive computer

try: #check to see if can access the Box folder
    cloud_folder = '/Users/zjcole/Box/file_drawer/'
except IOError: #in case can't connect to Box
    print('\n* ***** *\nNOT CONNECTED TO BOX\n\nThis will not be backed up....yet\n* ***** *\n')
    backup  = False #can't backup tp Box
else:
    cloud_folder = '/Users/zjcole/Box/file_drawer/'
    backup  = True #backup

# setup local file   
local_file = '%s%s' %(local_folder, local_filename)

# load the file
wb_local        = load_workbook(local_file)
ws_local        = wb_local["Sheet1"]
ws_local_rows   = wb_local["Sheet2"]

# would be nice to have a more prominent indicator of the beginning of this session
print('\n==============================\n\n==============================\n\n     ~---* NEW SESH *---~\n\n==============================\n\n==============================\n\n')

if ws_local['A1'].value == '':
    # input headers -- sheet 2
    ws_local['A1'].value = 'Date'
    ws_local['B1'].value = 'Time'
    ws_local['C1'].value = 'Project'
    ws_local['D1'].value = 'Project Status'
    ws_local['E1'].value = 'Task'
    ws_local['F1'].value = 'Task Status'
    ws_local['G1'].value = 'Task Time (secs)'
    ws_local['H1'].value = 'Notes'

    # input headers -- sheet 1
    ws_local_rows['A1'].value = 'rows'
    ws_local_rows['A2'].value = 1 #number of rows in the cloud
    ws_local_rows['B2'].value = 'Cloud'

    # save it
    wb_local.save(local_file)
    wb_local.close()

# thymer scripts
open_thymer  = 'open -a Thyme'
start_thymer = 'osascript -e \'tell app "Thyme" to start\''
stop_thymer  = 'osascript -e \'tell app "Thyme" to stop\''
close_thymer = 'osascript -e \'quit app "Thyme"\''

# date/time parameters
date = datetime.today().strftime('%m/%d/%Y')
time = datetime.today().strftime('%-H:%M')


#====================
# Setup Cloud Backup
#====================
# if able to backup to Box, compare everything in the Box (cloud) and local files
# if there is anything in the local file that is not in the Box file, upload to Box
# in the Box file, if oen fo the projects or tasks is complete, mark all of the rest of that particular task or project
# copy everything down from the cloud into the local file
if backup:

    #========================
    # Initialize Backup File
    #========================

    # open the workbook
    cloud_filename  = 'zptz_cloud.xlsx'
    cloud_file      = '%s%s' %(cloud_folder, cloud_filename)
    wb_cloud        = load_workbook(cloud_file)
    ws_cloud        = wb_cloud["Sheet1"] #Data
    ws_cloud_rows   = wb_cloud["Sheet2"] #RowCount

    if ws_cloud['A1'].value == '': #this should essentially mean that it is the first time using the script.
        # input headers -- sheet 1
        ws_cloud['A1'].value = 'Date'
        ws_cloud['B1'].value = 'Time'
        ws_cloud['C1'].value = 'Project'
        ws_cloud['D1'].value = 'Project Status'
        ws_cloud['E1'].value = 'Task'
        ws_cloud['F1'].value = 'Task Status'
        ws_cloud['G1'].value = 'Task Time (secs)'
        ws_cloud['H1'].value = 'Notes'

        # input headers - sheet 2
        ws_cloud_rows['A1'].value = 'rows'
        ws_cloud_rows['A2'].value = 1 #number of rows on the active local computer
        ws_cloud_rows['A3'].value = 1 #number of rows on the inactive local computer
        ws_cloud_rows['B2'].value = computer_name #logging computer name to keep track of which row number being used
        ws_cloud_rows['B3'].value = other_computer #^

        # save it and close it
        wb_cloud.save(cloud_file)
        wb_cloud.close()


    #================================================================
    # Push New Local Info to the Cloud (if necessary) and Vice Versa
    #================================================================
    # go through the local file info and save it to the cloud..
    # also will go through cloud and save that to local file..

    # activate for loop
    wb_local
    ws_local
    ws_local_rows

    wb_cloud
    ws_cloud
    ws_cloud_rows
    
    # setup max row
    mrow        = int(ws_local.max_row) #local file - max row plus one for recording
    mrow_cloud  = int(ws_cloud.max_row) #cloud file - max row plus one for recording

    cloud_rows_local = ws_local_rows['A2'].value #number of cloud rows in the local file last time it was saved

    if ws_cloud_rows['B2'].value == computer_name:
        local_rows_cloud = ws_cloud_rows['A2'].value #number of local rows in the cloud file last time it was saved
    else: #the other computer was backed up last time..
        local_rows_cloud = ws_cloud_rows['A3'].value #^^^

    # difference between previous number of rows, and current number of rows
    local_row_diff = int(mrow) - local_rows_cloud #did something change locally, but not get added to the cloud?
    cloud_row_diff = int(mrow_cloud) - cloud_rows_local #did something change in the cloud, but not locally?

    if (local_row_diff != 0) and (cloud_row_diff != 0): #if both changed (this would mean both were updated independently)
        # update  = 'both' #both of the worksheets need to be updated

        print('\n\n\nPushing to the cloud. . .\n\n\n')

        for d_row in range(1, local_row_diff): #cloud needs to be updated with local information
            diffrow     = str(int(mrow) - d_row)
            mrow_cloud  = str(int(ws_cloud.max_row) + 1) #cloud file - max row plus one for recording
            ws_cloud['A%s' %mrow_cloud].value = ws_local['A%s' %diffrow].value
            ws_cloud['B%s' %mrow_cloud].value = ws_local['B%s' %diffrow].value
            ws_cloud['C%s' %mrow_cloud].value = ws_local['C%s' %diffrow].value
            ws_cloud['D%s' %mrow_cloud].value = ws_local['D%s' %diffrow].value
            ws_cloud['E%s' %mrow_cloud].value = ws_local['E%s' %diffrow].value
            ws_cloud['F%s' %mrow_cloud].value = ws_local['F%s' %diffrow].value
            ws_cloud['G%s' %mrow_cloud].value = ws_local['G%s' %diffrow].value
            ws_cloud['H%s' %mrow_cloud].value = ws_local['H%s' %diffrow].value

            # save it
            wb_local.save(local_file)
            wb_cloud.save(cloud_file)

        print('\n\n\nPrecipitating from cloud. . .\n\n\n')

        for d_row in range(1, cloud_row_diff): #local needs to be updated with cloud information
            diffrow = str(int(mrow_cloud) - d_row - local_row_diff)
            mrow    = str(int(ws_local.max_row) + 1) #local file - max row plus one for recording
            ws_local['A%s' %mrow].value = ws_cloud['A%s' %diffrow].value
            ws_local['B%s' %mrow].value = ws_cloud['B%s' %diffrow].value
            ws_local['C%s' %mrow].value = ws_cloud['C%s' %diffrow].value
            ws_local['D%s' %mrow].value = ws_cloud['D%s' %diffrow].value
            ws_local['E%s' %mrow].value = ws_cloud['E%s' %diffrow].value
            ws_local['F%s' %mrow].value = ws_cloud['F%s' %diffrow].value
            ws_local['G%s' %mrow].value = ws_cloud['G%s' %diffrow].value
            ws_local['H%s' %mrow].value = ws_cloud['H%s' %diffrow].value

            # save it
            wb_local.save(local_file)
            wb_cloud.save(cloud_file)

    elif (local_row_diff != 0) and (cloud_row_diff == 0): #just the local sheet needs to be updated with the cloud information

        print('\n\n\nPrecipitating from cloud. . .\n\n\n')

        for d_row in range(1, cloud_row_diff): #local needs to be updated with cloud information
            diffrow = str(int(mrow_cloud) - d_row)
            mrow    = str(int(ws_local.max_row) + 1) #local file - max row plus one for recording
            ws_local['A%s' %mrow].value = ws_cloud['A%s' %diffrow].value
            ws_local['B%s' %mrow].value = ws_cloud['B%s' %diffrow].value
            ws_local['C%s' %mrow].value = ws_cloud['C%s' %diffrow].value
            ws_local['D%s' %mrow].value = ws_cloud['D%s' %diffrow].value
            ws_local['E%s' %mrow].value = ws_cloud['E%s' %diffrow].value
            ws_local['F%s' %mrow].value = ws_cloud['F%s' %diffrow].value
            ws_local['G%s' %mrow].value = ws_cloud['G%s' %diffrow].value
            ws_local['H%s' %mrow].value = ws_cloud['H%s' %diffrow].value

            # save it
            wb_local.save(local_file)
            wb_cloud.save(cloud_file)

    elif (local_row_diff == 0) and (cloud_row_diff != 0): #just the cloud sheet needs to be updated with the local information

        print('\n\n\nEvaporating to cloud. . . \n\n\n')

        for d_row in range(1, cloud_row_diff): #cloud needs to be updated with local information
            diffrow = str(int(mrow) - d_row)
            mrow_cloud  = str(int(ws_cloud.max_row) + 1) #cloud file - max row plus one for recording
            ws_cloud['A%s' %mrow_cloud].value = ws_local['A%s' %diffrow].value
            ws_cloud['B%s' %mrow_cloud].value = ws_local['B%s' %diffrow].value
            ws_cloud['C%s' %mrow_cloud].value = ws_local['C%s' %diffrow].value
            ws_cloud['D%s' %mrow_cloud].value = ws_local['D%s' %diffrow].value
            ws_cloud['E%s' %mrow_cloud].value = ws_local['E%s' %diffrow].value
            ws_cloud['F%s' %mrow_cloud].value = ws_local['F%s' %diffrow].value
            ws_cloud['G%s' %mrow_cloud].value = ws_local['G%s' %diffrow].value
            ws_cloud['H%s' %mrow_cloud].value = ws_local['H%s' %diffrow].value

            # save it
            wb_local.save(local_file)
            wb_cloud.save(cloud_file)

    elif (local_row_diff == 0) and (cloud_row_diff == 0): #no changes since last time
        print('\nNo need to backup. The cloud file is the same as the local file.')

    else:
        print('WARNING: Impossible changes option...')

    # save it
    wb_local.save(local_file)
    wb_local.close()
    wb_cloud.save(cloud_file)
    wb_cloud.close()


    #==================================================
    # Update the Cloud with Active and Completed Tasks
    #==================================================

    # initialize the files
    wb_cloud
    ws_cloud
    mrow_cloud  = str(int(ws_cloud.max_row) + 1)
    zmrow_cloud = range(2, int(mrow_cloud)) #need to define for the loop

    for zrow in zmrow_cloud:
        # work from the bottom of the list up
        # update for every loop iteration
        mrow_cloud  = str(int(ws_cloud.max_row) + 1)
        zrowx = str((int(mrow_cloud) + 1) - zrow) #add one because the minimum zrow value is 2; subtract zrow so works from the end of the file, up to the top

        # iteration variables
        # set these because it is just easier to type and less confusing
        project_update          = ws_cloud['C%s' %zrowx].value #project
        task_update             = ws_cloud['E%s' %zrowx].value #task
        project_update_status   = ws_cloud['D%s' %zrowx].value #project status
        task_update_status      = ws_cloud['F%s' %zrowx].value #task status

        # if the project is closed, close out everything
        if project_update_status == 'xx':
            for zrow_sub in zmrow_cloud: #cycle through the cloud file to update everything that matches the initial update values (above)
                project_update_sub          = ws_cloud['C%s' %zrow_sub].value
                project_update_status_sub   = ws_cloud['D%s' %zrow_sub].value
                task_update_status_sub      = ws_cloud['F%s' %zrow_sub].value       
                if project_update_sub == project_update: #if the project info is the same, close it down
                    project_update_status_sub   = 'xx'
                    task_update_status_sub      = 'xx'

        else: #if the project is still open, but the task is complete
            if task_update_status == 'xx':
                for zrow_sub in zmrow_cloud:
                    project_update_sub  = ws_cloud['C%s' %zrow_sub].value
                    task_update_sub     = ws_cloud['E%s' %zrow_sub].value
                    if project_update_sub == project_update: #if project is the same
                        if task_update_sub == task_update: #if task is the same
                            task_update_status_sub  = 'xx' #mark the task as complete
    
    # save it!
    wb_cloud.save(cloud_file)
    wb_cloud.close()
   
    # the cloud is now up to date
    print('\n\n\nThe cloud is updated.\n\n\n')

    # pull all from the cloud
    print('\n\n\nPulling from cloud. . . \n\n\n')

    
    #================================
    # Pull Everything from the Cloud
    #================================

    # re-open files
    wb_local
    ws_local
    wb_cloud
    ws_cloud

    # establish max row
    mrow_cloud  = str(int(ws_cloud.max_row))
    zmrow_cloud = range(1, int(mrow_cloud))

    # iterate through Box file and save everything to the local file, so they're identical
    for zrow in zmrow_cloud:
        # log data
        ws_local['A%s' %zrow].value = ws_cloud['A%s' %zrow].value #date
        ws_local['B%s' %zrow].value = ws_cloud['B%s' %zrow].value #time
        ws_local['C%s' %zrow].value = ws_cloud['C%s' %zrow].value #proj
        ws_local['D%s' %zrow].value = ws_cloud['D%s' %zrow].value #proj_status
        ws_local['E%s' %zrow].value = ws_cloud['E%s' %zrow].value #task
        ws_local['F%s' %zrow].value = ws_cloud['F%s' %zrow].value #task_status
        ws_local['G%s' %zrow].value = ws_cloud['G%s' %zrow].value #times
        ws_local['H%s' %zrow].value = ws_cloud['H%s' %zrow].value #notes
        
    # save logged data
    wb_local.save(local_file)
    wb_local.close()
    wb_cloud.close()

    # Data Pulled from the Cloud
    print('\n\n\nThe system is up to date!\n\n\n')

else: #if cannot backup to the cloud for some reason
    print('\n\n\n Not connected to the cloud....\n\n\n')
    

#============================
# Setup Interface Parameters
#============================

# window
SetOptions(background_color = 'black', element_background_color = 'black',
           text_color = 'white', text_element_background_color = 'black',
           element_text_color = 'white', input_elements_background_color = 'black',
           input_text_color = 'white')


#=====================
# Run the Task Manger
#=====================

# runit
exe_loop = None #main task loop
ploop = None #to select a project
while exe_loop == None:

    #==================================
    # Initiate Task Manager Parameters
    #==================================
    # working mainly off of the local file, but also updating the cloud file

    # open files
    wb_local #opening up new wb at beginning to account for newtask loop
    ws_local
    wb_cloud
    ws_cloud

    # establish max row
    # need to do this at the beginning of each loop in case stuff is logged to the file
    mrow    = str(int(ws_local.max_row) + 1)  #local file max row
    zmrow   = range(2, int(mrow))
    mrow_cloud  = str(int(ws_cloud.max_row)) #cloud file max row
    zmrow_cloud = range(1, int(mrow_cloud))
    
    #====================================
    # List Active Projects
    #====================================

    #find all of the projects marked with active status ('oo')
    project_list = [] #put the project names for all of the active projects in a list
    for zrow in zmrow:
        if ws_local['D%s' %zrow].value == 'oo':
            if ws_local['C%s' %zrow].value not in project_list: #only want one copy of each project name in the list
                project_list.append(ws_local['C%s' %zrow].value)  

    #=====================
    # Selecting a Project
    #=====================

    if ploop == None: #if need to select a project
        #list the projects
        print('\nProjects:')
        for project in project_list:
            print('(%s) %s' %(project_list.index(project), project))

        proj_loop = None #loop through the project unless the 'New Project' option is selected
        while proj_loop == None:
            try: #if no project information in the file
                # input with a number
                # NOTE: If want to add a new project, enter a number larger than the number of projects listed
                proj = int(raw_input('What do you want to work on? (Number): '))
            except ValueError:
                continue
            else:
                if 0 <= proj <= (len(project_list) - 1): #checking to see if project number is in the list
                    proj = project_list[proj]
                else: #if project is not in the list
                    pjn_loop = None #adding a new project
                    while pjn_loop == None:
                        projyn = raw_input('New project? (y/n): ')
                        if (projyn == '') or (projyn == 'y'): #just leave it empty or 'y' to confirm response
                            proj = raw_input('Project Name: ')
                            project_list.append(proj)
                            pjn_loop = 1

                            # if a 'completed' project is being re-activated
                            # go through file and update all project status's to active
                            for zrow in zmrow:
                                if ws_local['C%s' %zrow].value == proj:
                                    ws_local['D%s' %zrow].value = 'oo'
                                    wb_local.save(local_file)

                        elif projyn == 'n': #if not a new project
                            continue
                        else:
                            pjn_loop = None

            # confirm that this is the project you want to work on
            pl_loop = None
            while pl_loop == None:
                proj_l = raw_input('%s? (y/n): ' %proj)
                if (proj_l == '') or (proj_l == 'y'):
                    proj_loop = 1
                    pl_loop = 1
                elif proj_l == 'n':
                    pl_loop = 1
                else:
                    pl_loop = None

    else: #restart the project loop
        ploop = None

    #================
    # List the Notes
    #================

    # collect project notes to list
    proj_task_list = []
    proj_datetime_list = []
    projnote_list = []
    for zrow in zmrow:
        if ws_local['C%s' %zrow].value == proj: #for all notes in the associated project
            projnote_cell = ws_local['H%s' %zrow].value
            proj_task_cell = ws_local['E%s' %zrow].value
            proj_datetime_cell = str('%s %s' %(ws_local['A%s' %zrow].value, ws_local['B%s' %zrow].value)) #combining date and time for presentation purposes
            proj_task_list.append(proj_task_cell)
            proj_datetime_list.append(proj_datetime_cell)
            projnote_list.append(projnote_cell)

    # print the notes
    project_notes = []
    if len(proj_task_list) > 0: #if there are notes in the project
        print('\n-------------------------\n-------------------------\n    START %s NOTES\n-------------------------\n-------------------------' %proj.upper())
        # get all the notes written out
        for projn in range(0, len(proj_datetime_list)):
            project_note = '\n->%s /-/ %s\n%s' %(proj_task_list[projn], proj_datetime_list[projn], projnote_list[projn])
            if project_note not in project_notes:
                project_notes.append(project_note) #makiing sure each note is unique
                print(project_note)
        print('\n-------------------------\n-------------------------\n    END %s NOTES\n-------------------------\n-------------------------\n' %proj.upper()) #listing the project in all caps

    #=================================
    # List Active and Completed Tasks
    #=================================

    # list tasks
    task_list = []
    task_list_complete = []
    for zrow in zmrow:
        if ws_local['C%s' %zrow].value == proj: #if associated with the concerned project
            if ws_local['F%s' %zrow].value == 'oo': #if the task is active
                if ws_local['E%s' %zrow].value not in task_list: #if it is unique
                    task_list.append(ws_local['E%s' %zrow].value) #add it to the active task list
            else: #if the task is inactive
                if ws_local['E%s' %zrow].value not in task_list_complete: #if it is unique
                    task_list_complete.append(ws_local['E%s' %zrow].value) #add it to the complete task list

    # list the tasks
    print('\nCompleted Tasks:') #completed task list
    for completed_tasks in task_list_complete:
        print('-x-> %s' %(completed_tasks))           
    print('\nTasks:') #active task list
    for tasks in task_list:
        print('(%s) %s' %(task_list.index(tasks), tasks))

    #=============
    # Select Task
    #=============

    # selecting a task
    task_loop = 'n'
    while task_loop == 'n':
        try:
            task = int(raw_input('\nTask: #')) #making sure there are tasks in the list
            #if want to add a new task, enter number that is larger than the number of tasks in the list
        except ValueError:
            continue
        else:
            if 0 <= task <= (len(task_list) - 1): #making sure the task number is in the list
                task = task_list[task] #select the task
            else: #if task number entered isn't in the task list
                tyn_loop = None
                while tyn_loop == None: #if adding a new task
                    taskyn = raw_input('New task? (y/n): ')
                    if (taskyn == '') or (taskyn == 'y'): #hit enter, or 'y'
                        task = raw_input('Task Name: ') #add the task name
                        task_list.append(task) #add it to the task
                        tyn_loop = 1

                        #if a 'completed' task is being re-activated
                        for zrow in zmrow:
                            if ws_local['E%s' %zrow].value == task: #if the concerned task is already in the file
                                if ws_local['C%s' %zrow].value == proj: #if the concerned project is associated (making sure it is the correct task being selected)
                                    if ws_local['D%s' %zrow].value == 'oo': #if the project is active
                                        ws_local['F%s' %zrow].value = 'oo' #mark the task as active
                                        wb_local.save(local_file) #save it!

                    elif taskyn == 'n': #if it isn't a new task
                        continue
                    else:
                        tyn_loop = None
        
        # confirm this is the task you want to work on
        tl_loop = None
        while tl_loop == None:
            task_l = raw_input('%s? (y/n): ' %task)
            if (task_l == '') or (task_l == 'y'):
                task_loop = 1
                tl_loop = 1
            elif task_l == 'n':
                tl_loop = 1
                continue
            else:
                tl_loop = None
            
    #=================
    # List Task Notes
    #=================

    # collect the task notes
    tasknote_list = []
    task_datetime_list = []
    for zrow in zmrow:
        if ws_local['C%s' %zrow].value == proj:
            if ws_local['E%s' %zrow].value == task:
                tasknote_cell = ws_local['H%s' %zrow].value
                task_datetime_cell = str('%s %s' %(ws_local['A%s' %zrow].value, ws_local['B%s' %zrow].value))
                tasknote_list.append(tasknote_cell)
                task_datetime_list.append(task_datetime_cell)

    # list the notes
    task_notes = []
    if len(tasknote_list) > 0:
        print('\n->%s\n' %task.upper()) #print the notes associated with each task
        # get all the notes written out
        for taskn in range(0, len(task_datetime_list)):
            task_note = str('\n%s\n%s\n\n' %(task_datetime_list[taskn], tasknote_list[taskn]))
            if task_note not in task_notes: #make sure it is unique
                task_notes.append(task_note)
                print(task_note)

    #=====================
    # timing calculations
    #=====================
    # will collect timing in seconds, then mathematically workout the hours/mins
    proj_s = 0
    proj_time = 0
    proj_hours = 0
    proj_mins = 0
    task_s = 0
    task_time = 0
    task_mins = 0
    task_hours = 0
    for zrow in zmrow:
        # timing for the project
        if ws_local['C%s' %zrow].value == proj:
            ps = ws_local['G%s' %zrow].value
            if ps != None:
                proj_s += int(ps)

            # timing for the task
            if ws_local['E%s' %zrow].value == task:
                ts = ws_local['G%s' %zrow].value
                if ts != None:
                    task_s += int(ts)
    
    # calculate project timing
    if proj_s > 0:
        proj_hours = proj_s/3600
        proj_mins = (proj_s%3600)/60

        proj_time = '{h}h {m}m'.format(h = proj_hours, m = proj_mins)

    # calculate task timing
    if task_s > 0:
        task_hours = task_s/3600
        task_mins = (task_s%3600)/60
        task_time = '{h}h {m}m'.format(h = task_hours, m = task_mins)

    #====================
    # Activate Interface
    #====================

    # task timer - start
    task_start = datetime.now()
    
    #startup Thymer (operates in the top bar)
    os.system(close_thymer)
    os.system(open_thymer)
    os.system(start_thymer)

    # window
    p_window = sg.Window(str('%s: %s' %(date, time)), resizable = True, disable_close = True, finalize = True)
    p_layout = [[sg.Frame(layout = [[sg.Text('%s' %proj), sg.Text('%s' %proj_time)], [sg.Radio('In Progress', "projp", key = 'projip', default = True), sg.Radio('Completed', "projp")]], title = 'Project', relief = sg.RELIEF_SUNKEN)], [sg.Frame(layout = [[sg.Text('%s' %task), sg.Text('%s' %task_time)], [sg.Radio('In Progress', "taskp", key = 'taskip', default = True), sg.Radio('Completed', "taskp")]], title = 'Task', relief = sg.RELIEF_SUNKEN)], [sg.Multiline(size = (100, 8), key = 'notes', autoscroll = True, default_text = '----------\n')], [sg.CloseButton('Switch Project'), sg.CloseButton('Switch Task'), sg.CloseButton('Dunzo')]]

    #setup responses
    p_event, p_values = p_window.Layout(p_layout).Read()

    # task timer - end
    task_end = datetime.now()
    timex    = relativedelta(task_end, task_start)

    # timing calculations
    timeh = '{h}'.format(h = timex.hours)
    timem = '{m}'.format(m = timex.minutes)
    times = '{s}'.format(s = timex.seconds)

    timeh = int(timeh)*3600
    timem = int(timem)*60

    times = int(times)+timeh+timem

    os.system(stop_thymer)
    os.system(close_thymer)

    #===============
    # Log Responses
    #===============
    # input values

    # keep project active
    if p_values['projip'] == True:
        proj_status = 'oo'
    else: #if project is marked as completed
        #this will change everythig with that project name to completed
        proj_status = 'xx'
        #log it local and to the cloud
        for zrow in zmrow:
            if ws_local['C%s' %zrow].value == proj:
                ws_local['D%s' %zrow].value = proj_status #project status
                ws_local['F%s' %zrow].value = proj_status #task status
        for zrow in zmrow_cloud:
            if ws_cloud['C%s' %zrow].value == proj:
                ws_cloud['D%s' %zrow].value = proj_status #project status
                ws_cloud['F%s' %zrow].value = proj_status #task status

    # keep task active
    if p_values['taskip'] == True:
        task_status = 'oo'

    else: #if the task is marked as complete
        task_status = 'xx'
        for zrow in zmrow:
            #log it local and to the cloud
            if ws_local['C%s' %zrow].value == proj:
                if ws_local['E%s' %zrow].value == task:
                    ws_local['F%s' %zrow].value = task_status
        for zrow in zmrow_cloud:
            if ws_cloud['C%s' %zrow].value == proj:
                if ws_cloud['E%s' %zrow].value == task:
                    ws_cloud['F%s' %zrow].value = task_status
    
    # save the notes
    notes = p_values['notes']

    # log local
    mrow = str(int(ws_local.max_row) + 1) #reset every loop iteration so able to post to end of file

    ws_local['A%s' %mrow].value = date
    ws_local['B%s' %mrow].value = time
    ws_local['C%s' %mrow].value = proj
    ws_local['D%s' %mrow].value = proj_status
    ws_local['E%s' %mrow].value = task
    ws_local['F%s' %mrow].value = task_status
    ws_local['G%s' %mrow].value = times
    ws_local['H%s' %mrow].value = notes

    #close it out
    wb_local.save(local_file)
    wb_local.close()


    #============
    # Log Backup
    #============

    # back it all up to the cloud by adding the missing input logs
    # this whole thing might seem a little weird, but if we are backing up, the backup file will actually be the working documents file
    if backup:
        print('\n-----\nEvaporating to the cloud. . . \n\n\n')

        # just copy everything over to the cloud
        wb_cloud
        wb_local
        ws_cloud
        ws_local

        # max rows
        mrow = int(ws_local.max_row) + 1
        zmrow = range(2, mrow)

        # copy the local information directly to the cloud
        for zrow in zmrow:
            ws_cloud['A%s' %zrow].value = ws_local['A%s' %zrow].value #date
            ws_cloud['B%s' %zrow].value = ws_local['B%s' %zrow].value #time
            ws_cloud['C%s' %zrow].value = ws_local['C%s' %zrow].value #project
            ws_cloud['D%s' %zrow].value = ws_local['D%s' %zrow].value #project status
            ws_cloud['E%s' %zrow].value = ws_local['E%s' %zrow].value #task
            ws_cloud['F%s' %zrow].value = ws_local['F%s' %zrow].value #task status
            ws_cloud['G%s' %zrow].value = ws_local['G%s' %zrow].value #time
            ws_cloud['H%s' %zrow].value = ws_local['H%s' %zrow].value #notes

        # save it!
        wb_local.save(local_file)
        wb_local.close()
        wb_cloud.save(cloud_file)
        wb_cloud.close()

        # now calculate max rows
        # just copy everything over to the cloud
        wb_local
        ws_local
        ws_local_rows

        wb_cloud
        ws_cloud
        ws_cloud_rows

        # max rows
        mrow = int(ws_local.max_row)
        mrow_cloud = int(ws_cloud.max_row)

        # calculate the number of rows that are different
        if ws_cloud_rows['B2'].value == computer_name: #if the local computer was updated last time
            local_inactive_rows_local = ws_cloud_rows['A3'].value #number of local rows in the inactive document last time it was updated
        else: #if a different computer was updated last time
            local_inactive_rows_local = ws_cloud_rows['A2'].value #number of local rows in the active document last time it was updated

        ws_local_rows['A2'].value = zrow #number of rows in the cloud
        
        ws_cloud_rows['A2'].value = zrow #number of rows on the active local computer
        ws_cloud_rows['A3'].value = local_inactive_rows_local #number of rows on the inactive local computer
        ws_cloud_rows['B2'].value = computer_name
        ws_cloud_rows['B3'].value = other_computer

        #close it out
        wb_local.save(local_file)
        wb_local.close()
        wb_cloud.save(cloud_file)
        wb_cloud.close()

    else: #if not able to backup to the cloud
        #re-establish max row
        mrow = str(int(ws_local.max_row) + 1) 

        # log everything
        ws_local['A%s' %mrow].value = date
        ws_local['B%s' %mrow].value = time
        ws_local['C%s' %mrow].value = proj
        ws_local['D%s' %mrow].value = proj_status
        ws_local['E%s' %mrow].value = task
        ws_local['F%s' %mrow].value = task_status
        ws_local['G%s' %mrow].value = times
        ws_local['H%s' %mrow].value = notes

        #save logged data, close it down
        wb_local.save(local_file)
        wb_local.close()
        
        print('\n-----\nCould not backup to Box... will do next time (if connected).\n')

    # loop action
    if p_event == 'Dunzo': #if done working for now
        exe_loop = 1
    elif p_event == 'Switch Project': #if interested in switching the project
        continue
    elif p_event == 'Switch Task': #if interested in switching the task
        ploop = 1
        continue

# logout
exit()