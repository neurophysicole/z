#!/Library/Frameworks/Python.framework/Versions/2.7/bin/python
from pyexcel_xlsx import save_data
import pyexcel as pe
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import worksheet
import datetime
from dateutil.relativedelta import relativedelta
from datetime import datetime, date, timedelta
import os
import sys
from collections import OrderedDict
import pandas as pd
from pandas import read_excel
from pyexcel._compact import OrderedDict
import xlsxwriter
import PySimpleGUI27 as sg
from PySimpleGUI27 import SetOptions

# file parameters
try:
    folder   = '/Users/zcole/Box/file_drawer/'
except IOError:
    print('\n* ***** *\nNOT CONNECTED TO BOX\n* ***** *\n')
else:
    docfile = str('/Users/zcole/Documents/file_drawer/temp/zptz.xlsx')
    boxfile = str('/Users/zcole/Box/file_drawer/zptz.xlsx')

    # activate project wb
    wbd = load_workbook(docfile)
    wsd = wbd.active
    wbb = load_workbook(boxfile)
    wsb = wbb.active

    # rows and columns
    mdrow = str(int(wsd.max_row) + 1) 
    zmdrow = range(2, int(mdrow))
    mbrow = str(int(wsb.max_row) + 1)
    zmbrow = range(2, int(mbrow))
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    # copying zptz doc info to zptz Box
    for zdrow in zmdrow:
        for column in columns:
            wsb['%s%s' %(column, mbrow)].value = wsd['%s%s' %(column, zdrow)].value
        print('\nCopied Project: %s, Task: %s\n-- Note: %s\n' %(wsd['C%s' %zdrow].value, wsd['E%s' %zdrow].value, wsd['H%s' %zdrow].value))
        mbrow = str(int(mbrow) + 1)

    print('\n-----\nThe system is successfully backed up!')

    # close out excel files
    wbb.save(boxfile)
    wbb.close()
    wbd.close()

    # refresh the zptz Docs file
    os.system('rm docfile')
    os.system('touch docfile')

    print('\nThe backup system is refreshed.\n-----')

# logout
exit()